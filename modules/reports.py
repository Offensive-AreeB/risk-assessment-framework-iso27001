import streamlit as st
import pandas as pd
import os
from datetime import datetime, date as date_type
from database.db import execute_query
from io import BytesIO

# ── ReportLab imports ────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from utils.rbac import has_permission
from utils.audit import log_action
from reportlab.platypus import (
    BaseDocTemplate, PageTemplate, Frame, NextPageTemplate,
    Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether, HRFlowable
)
from reportlab.platypus.flowables import Flowable
from reportlab.lib.units import inch, cm
from reportlab.pdfgen import canvas as pdf_canvas

# ── openpyxl imports ─────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
REPORTS_DIR = "reports/generated"
os.makedirs(REPORTS_DIR, exist_ok=True)

# ── Brand Palette ─────────────────────────────────────────────────────────────
NAVY        = colors.HexColor("#0D1B3E")   # primary dark
DARK_BLUE   = colors.HexColor("#1A3A6B")
MID_BLUE    = colors.HexColor("#2E6DB4")
LIGHT_BLUE  = colors.HexColor("#D6E4F7")
LIGHTEST    = colors.HexColor("#F2F6FC")
GREY_DARK   = colors.HexColor("#4A4A4A")
GREY_MID    = colors.HexColor("#8C8C8C")
GREY_LIGHT  = colors.HexColor("#E8E8E8")
WHITE       = colors.white
BLACK       = colors.black

# Severity colours
C_CRITICAL  = colors.HexColor("#C0392B")
C_HIGH      = colors.HexColor("#E67E22")
C_MEDIUM    = colors.HexColor("#F1C40F")
C_LOW       = colors.HexColor("#27AE60")

# Severity colour map
LEVEL_COLOR = {
    "Critical":  C_CRITICAL,
    "High":      C_HIGH,
    "Medium":    C_MEDIUM,
    "Low":       C_LOW,
}

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def score_color(score):
    if score >= 17: return C_CRITICAL
    if score >= 10: return C_HIGH
    if score >= 5:  return C_MEDIUM
    return C_LOW


def is_overdue(status, target_date):
    if status in ("Implemented", "Accepted", "Closed"):
        return False
    try:
        d = date_type.fromisoformat(str(target_date))
        return d < date_type.today()
    except Exception:
        return False


def fetch_report_data(scope, f_asset, f_level, f_status):
    risks = execute_query(
        """SELECT r.id, r.risk_title, a.asset_name, r.threat, r.vulnerability,
                  r.existing_controls, r.likelihood, r.impact, r.risk_score,
                  r.risk_level, r.risk_owner, r.status
           FROM risks r JOIN assets a ON r.asset_id = a.id""", fetch_all=True)

    treatments = execute_query(
        """SELECT t.id, r.risk_title, t.treatment_option, t.treatment_description,
                  t.treatment_owner, t.target_date, t.treatment_status,
                  t.residual_likelihood, t.residual_impact,
                  t.residual_score, t.residual_risk_level, t.risk_id,
                  r.risk_score as inherent_score, r.risk_level as inherent_level
           FROM risk_treatments t JOIN risks r ON t.risk_id = r.id""", fetch_all=True)

    mappings = execute_query(
        """SELECT m.id, r.risk_title, c.control_id, c.control_name, c.control_category,
                  m.applicability, m.justification, m.implementation_status,
                  m.implementation_notes
           FROM risk_control_mapping m
           JOIN risks r ON m.risk_id = r.id
           JOIN iso_controls c ON m.control_id = c.control_id""", fetch_all=True)

    assets  = execute_query("SELECT * FROM assets",     fetch_all=True)
    controls = execute_query("SELECT * FROM iso_controls", fetch_all=True)

    filtered_risks = [dict(r) for r in risks]
    if f_asset  != "All": filtered_risks = [r for r in filtered_risks if r["asset_name"] == f_asset]
    if f_level  != "All": filtered_risks = [r for r in filtered_risks if r["risk_level"]  == f_level]
    if f_status != "All": filtered_risks = [r for r in filtered_risks if r["status"]       == f_status]

    valid_titles = {r["risk_title"] for r in filtered_risks}
    f_treatments = [dict(t) for t in treatments if t["risk_title"] in valid_titles]
    f_mappings   = [dict(m) for m in mappings   if m["risk_title"] in valid_titles]

    return (filtered_risks, f_treatments, f_mappings,
            [dict(a) for a in assets], [dict(c) for c in controls])


def compute_kpis(risks, treatments, mappings, assets):
    crit  = sum(1 for r in risks if r["risk_level"] == "Critical")
    high  = sum(1 for r in risks if r["risk_level"] == "High")
    med   = sum(1 for r in risks if r["risk_level"] == "Medium")
    low   = sum(1 for r in risks if r["risk_level"] == "Low")
    app_c = len({m["control_id"] for m in mappings if m["applicability"] == "Applicable"})
    treated = {t["risk_title"] for t in treatments}
    no_treat = sum(1 for r in risks if r["risk_title"] not in treated)
    open_t   = sum(1 for t in treatments if t["treatment_status"] in ("Planned", "In Progress"))
    overdue  = sum(1 for t in treatments if is_overdue(t["treatment_status"], t["target_date"]))
    return dict(
        total_assets=len(assets), total_risks=len(risks),
        critical=crit, high=high, medium=med, low=low,
        applicable_controls=app_c, no_treatment=no_treat,
        open_treatments=open_t, overdue_treatments=overdue
    )


# ─────────────────────────────────────────────────────────────────────────────
# Style helpers
# ─────────────────────────────────────────────────────────────────────────────
def get_styles():
    base = getSampleStyleSheet()
    styles = {}

    def add(name, **kw):
        styles[name] = ParagraphStyle(name=name, **kw)

    add("Cover_Title",
        fontName="Helvetica-Bold", fontSize=28, textColor=WHITE,
        leading=34, spaceAfter=8, alignment=TA_LEFT)
    add("Cover_Sub",
        fontName="Helvetica", fontSize=14, textColor=LIGHT_BLUE,
        leading=18, spaceAfter=4, alignment=TA_LEFT)
    add("Cover_Label",
        fontName="Helvetica-Bold", fontSize=9, textColor=GREY_MID,
        spaceAfter=2, alignment=TA_LEFT, letterSpacing=1)
    add("Cover_Value",
        fontName="Helvetica", fontSize=11, textColor=WHITE,
        spaceAfter=6, alignment=TA_LEFT)
    add("Cover_Conf",
        fontName="Helvetica-Bold", fontSize=8, textColor=GREY_MID,
        alignment=TA_LEFT, letterSpacing=2)

    add("Sect_Title",
        fontName="Helvetica-Bold", fontSize=16, textColor=NAVY,
        spaceAfter=6, spaceBefore=12, leading=20)
    add("Sub_Title",
        fontName="Helvetica-Bold", fontSize=12, textColor=DARK_BLUE,
        spaceAfter=4, spaceBefore=8, leading=15)
    add("Body",
        fontName="Helvetica", fontSize=9, textColor=GREY_DARK,
        leading=13, spaceAfter=4)
    add("Body_Bold",
        fontName="Helvetica-Bold", fontSize=9, textColor=GREY_DARK,
        leading=13)
    add("Caption",
        fontName="Helvetica", fontSize=8, textColor=GREY_MID,
        leading=11, alignment=TA_CENTER)
    add("Table_Header",
        fontName="Helvetica-Bold", fontSize=8, textColor=WHITE,
        alignment=TA_CENTER, leading=10)
    add("Table_Cell",
        fontName="Helvetica", fontSize=8, textColor=GREY_DARK,
        leading=11, wordWrap="CJK")
    add("Table_Cell_Bold",
        fontName="Helvetica-Bold", fontSize=8, textColor=GREY_DARK,
        leading=11)
    add("KPI_Value",
        fontName="Helvetica-Bold", fontSize=22, textColor=NAVY,
        alignment=TA_CENTER, leading=26)
    add("KPI_Label",
        fontName="Helvetica", fontSize=8, textColor=GREY_MID,
        alignment=TA_CENTER, leading=10)
    add("Posture",
        fontName="Helvetica", fontSize=10, textColor=GREY_DARK,
        leading=15, spaceAfter=6, borderPadding=(8, 8, 8, 8))

    return styles


# ─────────────────────────────────────────────────────────────────────────────
# Page template — header / footer with page numbers
# ─────────────────────────────────────────────────────────────────────────────
class NumberedCanvas(pdf_canvas.Canvas):
    """ReportLab 5-compatible Page X of Y footer canvas."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._page_states = []

    def showPage(self):
        self._page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total = len(self._page_states)
        for page_num, state in enumerate(self._page_states, start=1):
            self.__dict__.update(state)
            self._draw_decoration(page_num, total)
            super().showPage()
        super().save()

    def _draw_decoration(self, page_num, total_pages):
        w, h = landscape(A4)

        # Skip cover page (page 1)
        if page_num == 1:
            return

        # ── Header bar ───────────────────────────────────────────────────
        self.setFillColor(NAVY)
        self.rect(0, h - 30, w, 30, fill=1, stroke=0)
        self.setFont("Helvetica-Bold", 9)
        self.setFillColor(WHITE)
        self.drawString(15, h - 19, "Risk Assessment Framework")
        self.setFont("Helvetica", 8)
        self.setFillColor(LIGHT_BLUE)
        self.drawString(15, h - 28, "ISO/IEC 27001-Aligned GRC Assessment")

        # ── Footer bar ───────────────────────────────────────────────────
        self.setFillColor(NAVY)
        self.rect(0, 0, w, 22, fill=1, stroke=0)
        self.setFont("Helvetica", 7.5)
        self.setFillColor(GREY_MID)
        self.drawString(15, 7, "CONFIDENTIAL -- INTERNAL USE ONLY")
        self.setFillColor(WHITE)
        self.drawCentredString(w / 2, 7, "Risk Assessment Framework")
        self.drawRightString(w - 15, 7, f"Page {page_num} of {total_pages}")


# ─────────────────────────────────────────────────────────────────────────────
# Cover page template drawing
# ─────────────────────────────────────────────────────────────────────────────
def draw_cover(c, doc):
    scope = getattr(doc, "report_scope", "Report")
    gen_date = getattr(doc, "gen_date", "")
    pw, ph = landscape(A4)

    # 1. Background split
    c.setFillColor(NAVY)
    c.rect(0, 0, pw * 0.52, ph, fill=1, stroke=0)
    c.setFillColor(LIGHTEST)
    c.rect(pw * 0.52, 0, pw * 0.48, ph, fill=1, stroke=0)

    # Accent stripe
    c.setFillColor(MID_BLUE)
    c.rect(pw * 0.52 - 4, 0, 4, ph, fill=1, stroke=0)

    # 2. Minimal geometric visual (right side)
    cx, cy = pw * 0.76, ph * 0.45
    c.setStrokeColor(LIGHT_BLUE)
    c.setLineWidth(1)
    # Background grid motif
    grid_sz = 20
    for i in range(-4, 5):
        c.line(cx - 80, cy + i * grid_sz, cx + 80, cy + i * grid_sz)
        c.line(cx + i * grid_sz, cy - 80, cx + i * grid_sz, cy + 80)

    # Professional shield outline
    c.setStrokeColor(DARK_BLUE)
    c.setLineWidth(2.5)
    c.setLineJoin(1)
    p = c.beginPath()
    p.moveTo(cx, cy + 60)
    p.lineTo(cx - 40, cy + 30)
    p.lineTo(cx - 40, cy - 20)
    p.curveTo(cx - 40, cy - 60, cx, cy - 90, cx, cy - 90)
    p.curveTo(cx, cy - 90, cx + 40, cy - 60, cx + 40, cy - 20)
    p.lineTo(cx + 40, cy + 30)
    p.close()
    c.drawPath(p, fill=0, stroke=1)

    # Inner motif
    c.setFillColor(MID_BLUE)
    c.rect(cx - 15, cy - 15, 30, 30, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.circle(cx, cy, 6, fill=1, stroke=0)

    # 3. Typography Hierarchy (left side)
    left_x = 45

    # Small Top Label
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(LIGHT_BLUE)
    c.drawString(left_x, ph - 65, "INFORMATION SECURITY | GOVERNANCE, RISK & COMPLIANCE")

    # Main Title
    c.setFont("Helvetica-Bold", 32)
    c.setFillColor(WHITE)
    c.drawString(left_x, ph - 120, "RISK ASSESSMENT")
    c.drawString(left_x, ph - 160, "FRAMEWORK")

    # Subtitle
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(LIGHT_BLUE)
    c.drawString(left_x, ph - 200, "ISO/IEC 27001:2022-Aligned GRC Risk Assessment")

    c.setFont("Helvetica", 10)
    c.setFillColor(WHITE)
    c.drawString(left_x, ph - 220, "RISK ASSESSMENT & COMPLIANCE REPORT")

    # Divider
    c.setStrokeColor(MID_BLUE)
    c.setLineWidth(1)
    c.line(left_x, ph - 245, pw * 0.45, ph - 245)

    # Report Information Fields
    y = ph - 290
    fields = [
        ("Prepared by", "Areeb Amjad Khan"),
        ("Organization", "Sample Organization"),
        ("Project ID", "ZYNVEX-CERT-0666"),
        ("Report Type", scope),
        ("Report Version", "1.0"),
        ("Generated", gen_date),
    ]
    for label, val in fields:
        c.setFont("Helvetica-Bold", 8)
        c.setFillColor(LIGHT_BLUE)
        c.drawString(left_x, y, label.upper())
        c.setFont("Helvetica", 10)
        c.setFillColor(WHITE)
        c.drawString(left_x, y - 12, val)
        y -= 36

    # Bottom Confidentiality
    c.setFont("Helvetica-Bold", 8)
    c.setFillColor(GREY_MID)
    c.drawString(left_x, 30, "CONFIDENTIAL — INTERNAL USE ONLY")


# ─────────────────────────────────────────────────────────────────────────────
# Section header helper
# ─────────────────────────────────────────────────────────────────────────────
def section_header(title, styles):
    class SectionBar(Flowable):
        def draw(self):
            c = self.canv
            w = landscape(A4)[0] - 60  # page width minus margins
            c.setFillColor(NAVY)
            c.rect(0, 0, w, 24, fill=1, stroke=0)
            c.setFillColor(MID_BLUE)
            c.rect(0, 0, 5, 24, fill=1, stroke=0)
            c.setFont("Helvetica-Bold", 12)
            c.setFillColor(WHITE)
            c.drawString(12, 7, title.upper())
        def wrap(self, available_width, available_height):
            return available_width, 24
    return [SectionBar(), Spacer(1, 10)]


# ─────────────────────────────────────────────────────────────────────────────
# Styled table builder
# ─────────────────────────────────────────────────────────────────────────────
def build_table(data_rows, col_widths, styles_obj, alt_start=1):
    """
    data_rows[0] = header row (list of strings).
    Returns a ReportLab Table with professional styling.
    """
    sty = styles_obj
    cell_style = sty["Table_Cell"]
    hdr_style  = sty["Table_Header"]

    # Convert strings to Paragraphs for wrapping
    table_data = []
    for i, row in enumerate(data_rows):
        if i == 0:
            table_data.append([Paragraph(str(c), hdr_style) for c in row])
        else:
            table_data.append([Paragraph(str(c) if c is not None else "", cell_style) for c in row])

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    cmd = [
        # Header
        ("BACKGROUND",  (0, 0), (-1, 0),  NAVY),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
        ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
        ("ROWBACKGROUNDS", (0, alt_start), (-1, -1), [WHITE, LIGHTEST]),
        ("GRID",        (0, 0), (-1, -1), 0.3, GREY_LIGHT),
        ("LINEBELOW",   (0, 0), (-1, 0),  1.2, MID_BLUE),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(cmd))
    return t


def severity_cell_style(table, data_rows, col_index):
    """Apply severity background to cells in col_index based on level text."""
    cmds = []
    for i, row in enumerate(data_rows[1:], start=1):
        level = str(row[col_index]).strip()
        bg = LEVEL_COLOR.get(level)
        if bg:
            cmds.append(("BACKGROUND", (col_index, i), (col_index, i), bg))
            cmds.append(("TEXTCOLOR",  (col_index, i), (col_index, i), WHITE))
            cmds.append(("FONTNAME",   (col_index, i), (col_index, i), "Helvetica-Bold"))
    if cmds:
        table.setStyle(TableStyle(cmds))
    return table


# ─────────────────────────────────────────────────────────────────────────────
# KPI cards (mini-table)
# ─────────────────────────────────────────────────────────────────────────────
def build_kpi_cards(kpis, styles_obj):
    sty = styles_obj
    cards = [
        ("Total\nAssets",      str(kpis["total_assets"]),      MID_BLUE),
        ("Total\nRisks",       str(kpis["total_risks"]),        DARK_BLUE),
        ("Critical",           str(kpis["critical"]),           C_CRITICAL),
        ("High",               str(kpis["high"]),               C_HIGH),
        ("Medium",             str(kpis["medium"]),             C_MEDIUM),
        ("Low",                str(kpis["low"]),                C_LOW),
        ("Applicable\nControls", str(kpis["applicable_controls"]), MID_BLUE),
        ("Risks w/o\nTreatment", str(kpis["no_treatment"]),    C_HIGH),
        ("Open\nTreatments",   str(kpis["open_treatments"]),   DARK_BLUE),
        ("Overdue\nTreatments",str(kpis["overdue_treatments"]), C_CRITICAL),
    ]

    card_width  = (landscape(A4)[0] - 60) / len(cards)
    header_row  = []
    value_row   = []
    colour_cmds = []

    for idx, (label, value, colour) in enumerate(cards):
        header_row.append(Paragraph(label, ParagraphStyle(
            name=f"kl_{idx}", fontName="Helvetica", fontSize=7.5,
            textColor=WHITE, alignment=TA_CENTER, leading=10)))
        value_row.append(Paragraph(value, ParagraphStyle(
            name=f"kv_{idx}", fontName="Helvetica-Bold", fontSize=20,
            textColor=WHITE, alignment=TA_CENTER, leading=24)))
        colour_cmds += [
            ("BACKGROUND", (idx, 0), (idx, 0), colour),
            ("BACKGROUND", (idx, 1), (idx, 1), colour),
        ]

    t = Table([header_row, value_row], colWidths=[card_width] * len(cards))
    style_cmds = colour_cmds + [
        ("TOPPADDING",     (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
        ("LINEAFTER",      (0, 0), (-1, -1), 0.5, WHITE),
        ("ROWBACKGROUNDS", (0, 2), (-1, -1), [LIGHTEST]),
    ]
    t.setStyle(TableStyle(style_cmds))
    return t


# ─────────────────────────────────────────────────────────────────────────────
# Risk Distribution horizontal bar chart
# ─────────────────────────────────────────────────────────────────────────────
def build_risk_bar_chart(kpis, styles_obj):
    levels  = ["Critical", "High", "Medium", "Low"]
    counts  = [kpis["critical"], kpis["high"], kpis["medium"], kpis["low"]]
    clrs    = [C_CRITICAL, C_HIGH, C_MEDIUM, C_LOW]
    total   = max(sum(counts), 1)
    pw      = landscape(A4)[0] - 60
    bar_w   = pw * 0.55

    class BarChart(Flowable):
        def __init__(self):
            self._height = len(levels) * 28 + 10
        def wrap(self, aw, ah):
            return aw, self._height
        def draw(self):
            c = self.canv
            row_h = 28
            label_x = 0
            bar_x   = 80
            y = self._height - row_h
            for level, count, clr in zip(levels, counts, clrs):
                filled = int(bar_w * (count / total))
                # Background track
                c.setFillColor(GREY_LIGHT)
                c.roundRect(bar_x, y + 6, bar_w, 16, 3, fill=1, stroke=0)
                # Filled bar
                if filled > 6:
                    c.setFillColor(clr)
                    c.roundRect(bar_x, y + 6, filled, 16, 3, fill=1, stroke=0)
                # Label
                c.setFont("Helvetica-Bold", 9)
                c.setFillColor(GREY_DARK)
                c.drawString(label_x, y + 10, level)
                # Count badge
                c.setFont("Helvetica-Bold", 9)
                c.setFillColor(clr)
                c.drawString(bar_x + bar_w + 8, y + 10, str(count))
                y -= row_h

    return BarChart()


# ─────────────────────────────────────────────────────────────────────────────
# 5×5 Risk Matrix
# ─────────────────────────────────────────────────────────────────────────────
def build_risk_matrix(risks, styles_obj):
    sty = styles_obj
    x_labels = ["1\nRare", "2\nUnlikely", "3\nPossible", "4\nLikely", "5\nAlmost\nCertain"]
    y_labels = ["5\nSevere", "4\nMajor", "3\nModerate", "2\nMinor", "1\nInsignificant"]

    pw = landscape(A4)[0] - 60
    cell_w = (pw - 70) / 5

    # Header row: empty corner + likelihood labels
    hdr = [Paragraph("Impact \\ Likelihood", ParagraphStyle(
               name="mh", fontName="Helvetica-Bold", fontSize=8,
               textColor=WHITE, alignment=TA_CENTER, leading=10))]
    for xl in x_labels:
        hdr.append(Paragraph(xl, ParagraphStyle(
            name="mhx", fontName="Helvetica-Bold", fontSize=8,
            textColor=WHITE, alignment=TA_CENTER, leading=10)))

    table_data = [hdr]
    style_cmds = [
        ("BACKGROUND",    (0, 0), (-1, 0), NAVY),
        ("BACKGROUND",    (0, 0), (0, -1), NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0), WHITE),
        ("TEXTCOLOR",     (0, 0), (0, -1), WHITE),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",          (0, 0), (-1, -1), 0.5, WHITE),
        ("TOPPADDING",    (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
    ]

    for row_i, impact in enumerate(range(5, 0, -1), start=1):
        row = [Paragraph(y_labels[row_i - 1], ParagraphStyle(
            name=f"yl_{row_i}", fontName="Helvetica-Bold", fontSize=7.5,
            textColor=WHITE, alignment=TA_CENTER, leading=10))]
        for likelihood in range(1, 6):
            score = likelihood * impact
            bg    = score_color(score)
            count = sum(1 for r in risks
                        if r["likelihood"] == likelihood and r["impact"] == impact)
            cell_text = (
                f'<b>{score}</b>'
                f'<br/><font size="7">{count} risk{"s" if count != 1 else ""}</font>'
            )
            cell_p = Paragraph(cell_text, ParagraphStyle(
                name=f"mc_{row_i}_{likelihood}", fontName="Helvetica-Bold",
                fontSize=10, textColor=WHITE, alignment=TA_CENTER, leading=14))
            row.append(cell_p)
            col = likelihood   # 1-indexed
            style_cmds.append(("BACKGROUND", (col, row_i), (col, row_i), bg))
        table_data.append(row)

    col_widths = [70] + [cell_w] * 5
    t = Table(table_data, colWidths=col_widths, rowHeights=[36] + [48] * 5)
    t.setStyle(TableStyle(style_cmds))

    # Legend
    legend_items = [
        ("Low (1–4)",       C_LOW),
        ("Medium (5–9)",    C_MEDIUM),
        ("High (10–16)",    C_HIGH),
        ("Critical (17–25)",C_CRITICAL),
    ]
    leg_data  = [[Paragraph(l, ParagraphStyle(
                    name=f"leg_{idx}", fontName="Helvetica-Bold", fontSize=8,
                    textColor=WHITE, alignment=TA_CENTER))
                  for idx, (l, _) in enumerate(legend_items)]]
    leg_table = Table(leg_data, colWidths=[(pw / 4)] * 4)
    leg_cmds  = [("ALIGN", (0, 0), (-1, -1), "CENTER"),
                 ("TOPPADDING", (0, 0), (-1, -1), 5),
                 ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]
    for i, (_, c) in enumerate(legend_items):
        leg_cmds.append(("BACKGROUND", (i, 0), (i, 0), c))
    leg_table.setStyle(TableStyle(leg_cmds))

    return [t, Spacer(1, 6), leg_table]


# ─────────────────────────────────────────────────────────────────────────────
# Executive Summary
# ─────────────────────────────────────────────────────────────────────────────
def build_exec_summary(risks, treatments, mappings, assets, styles_obj):
    sty  = styles_obj
    kpis = compute_kpis(risks, treatments, mappings, assets)
    els  = []

    els += section_header("Executive Summary", styles_obj)
    els.append(build_kpi_cards(kpis, styles_obj))
    els.append(Spacer(1, 14))

    # Posture paragraph (dynamic)
    overdue_note = (f" There are {kpis['overdue_treatments']} overdue treatment"
                    f"{'s' if kpis['overdue_treatments'] != 1 else ''} requiring immediate attention."
                    if kpis["overdue_treatments"] > 0 else "")
    posture = (
        f"The current risk profile for Sample Organization contains "
        f"<b>{kpis['total_risks']}</b> identified risks across "
        f"<b>{kpis['total_assets']}</b> registered assets. "
        f"This includes <b>{kpis['critical']} Critical</b>, "
        f"<b>{kpis['high']} High</b>, "
        f"<b>{kpis['medium']} Medium</b>, and "
        f"<b>{kpis['low']} Low</b> risk items. "
        f"<b>{kpis['applicable_controls']}</b> ISO/IEC 27001:2022 Annex A controls "
        f"have been assessed as Applicable. "
        f"<b>{kpis['no_treatment']}</b> risk"
        f"{'s do' if kpis['no_treatment'] != 1 else ' does'} not yet have a "
        f"treatment plan, and <b>{kpis['open_treatments']}</b> treatment"
        f"{'s are' if kpis['open_treatments'] != 1 else ' is'} currently active.{overdue_note}"
    )
    els.append(Paragraph("<b>RISK POSTURE</b>", ParagraphStyle(
        name="posture_hdr", fontName="Helvetica-Bold", fontSize=9,
        textColor=DARK_BLUE, spaceBefore=4, spaceAfter=4)))
    els.append(HRFlowable(width="100%", thickness=1, color=LIGHT_BLUE,
                           spaceAfter=6))
    els.append(Paragraph(posture, sty["Posture"]))
    els.append(Spacer(1, 10))

    # Risk Distribution chart
    els.append(Paragraph("<b>RISK DISTRIBUTION</b>", ParagraphStyle(
        name="dist_hdr", fontName="Helvetica-Bold", fontSize=9,
        textColor=DARK_BLUE, spaceAfter=4)))
    els.append(HRFlowable(width="100%", thickness=1, color=LIGHT_BLUE,
                           spaceAfter=6))
    els.append(build_risk_bar_chart(kpis, styles_obj))
    els.append(PageBreak())
    return els


# ─────────────────────────────────────────────────────────────────────────────
# Risk Matrix section
# ─────────────────────────────────────────────────────────────────────────────
def build_matrix_section(risks, styles_obj):
    els = []
    els += section_header("5×5 Risk Matrix", styles_obj)
    els.append(Paragraph(
        "The matrix plots each risk by Likelihood (X-axis) and Impact (Y-axis). "
        "Cell values show the Risk Score and count of risks in that position.",
        styles_obj["Body"]))
    els.append(Spacer(1, 8))
    els += build_risk_matrix(risks, styles_obj)
    els.append(PageBreak())
    return els


# ─────────────────────────────────────────────────────────────────────────────
# Top Risks
# ─────────────────────────────────────────────────────────────────────────────
def build_top_risks(risks, styles_obj):
    sty = styles_obj
    els = []
    els += section_header("Top Risks", styles_obj)

    sorted_risks = sorted(risks, key=lambda r: r["risk_score"], reverse=True)[:10]
    pw = landscape(A4)[0] - 60
    headers = ["#", "Risk Title", "Asset", "Score", "Risk Level", "Status"]
    widths  = [22, pw*0.34, pw*0.18, 38, 65, 80]

    rows = [headers]
    for idx, r in enumerate(sorted_risks, 1):
        rows.append([
            str(idx),
            r["risk_title"],
            r["asset_name"],
            str(r["risk_score"]),
            r["risk_level"],
            r["status"],
        ])

    t = build_table(rows, widths, sty)
    t = severity_cell_style(t, rows, 4)
    els.append(KeepTogether(t))
    els.append(Spacer(1, 8))
    return els


# ─────────────────────────────────────────────────────────────────────────────
# Risk Register
# ─────────────────────────────────────────────────────────────────────────────
def build_risk_register(risks, styles_obj):
    sty = styles_obj
    els = []
    els += section_header("Risk Register", styles_obj)

    pw = landscape(A4)[0] - 60
    headers = ["Risk Title", "Asset", "Threat", "L", "I", "Score",
               "Risk Level", "Owner", "Status"]
    widths  = [pw*0.22, pw*0.12, pw*0.20, 20, 20, 30, 55, pw*0.12, 80]

    rows = [headers]
    for r in risks:
        rows.append([
            r["risk_title"],
            r["asset_name"],
            r["threat"] or "",
            str(r["likelihood"]),
            str(r["impact"]),
            str(r["risk_score"]),
            r["risk_level"],
            r["risk_owner"] or "",
            r["status"] or "",
        ])

    t = build_table(rows, widths, sty)
    t = severity_cell_style(t, rows, 6)
    els.append(t)
    els.append(PageBreak())
    return els


# ─────────────────────────────────────────────────────────────────────────────
# ISO / SoA section
# ─────────────────────────────────────────────────────────────────────────────
def build_soa_section(mappings, styles_obj):
    sty = styles_obj
    els = []
    els += section_header("ISO/IEC 27001:2022 — Statement of Applicability", styles_obj)

    # Build SoA dict
    soa = {}
    for m in mappings:
        cid = m["control_id"]
        if cid not in soa:
            soa[cid] = {"name": m["control_name"], "theme": m["control_category"],
                        "app": m["applicability"],
                        "stat": m["implementation_status"] or ""}
        else:
            if m["applicability"] == "Applicable":
                soa[cid]["app"] = "Applicable"

    # KPI mini row
    app_count  = sum(1 for v in soa.values() if v["app"] == "Applicable")
    napp_count = sum(1 for v in soa.values() if v["app"] == "Not Applicable")
    impl_count = sum(1 for v in soa.values() if "Implemented" in v["stat"]
                     and "Not" not in v["stat"])
    part_count = sum(1 for v in soa.values() if "Partially" in v["stat"])
    plan_count = sum(1 for v in soa.values() if "Planned" in v["stat"])

    mini_data = [[
        Paragraph(f"<b>{len(soa)}</b><br/>Total", ParagraphStyle(
            name="sc0", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
        Paragraph(f"<b>{app_count}</b><br/>Applicable", ParagraphStyle(
            name="sc1", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
        Paragraph(f"<b>{napp_count}</b><br/>Not Applicable", ParagraphStyle(
            name="sc2", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
        Paragraph(f"<b>{impl_count}</b><br/>Implemented", ParagraphStyle(
            name="sc3", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
        Paragraph(f"<b>{part_count}</b><br/>Partial", ParagraphStyle(
            name="sc4", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
        Paragraph(f"<b>{plan_count}</b><br/>Planned", ParagraphStyle(
            name="sc5", fontName="Helvetica", fontSize=8, textColor=WHITE,
            alignment=TA_CENTER, leading=11)),
    ]]
    pw = landscape(A4)[0] - 60
    mini_t = Table(mini_data, colWidths=[pw/6]*6)
    mini_t.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (0,0), DARK_BLUE),
        ("BACKGROUND",    (1,0), (1,0), MID_BLUE),
        ("BACKGROUND",    (2,0), (2,0), GREY_MID),
        ("BACKGROUND",    (3,0), (3,0), C_LOW),
        ("BACKGROUND",    (4,0), (4,0), C_MEDIUM),
        ("BACKGROUND",    (5,0), (5,0), MID_BLUE),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LINEAFTER",     (0,0), (-1,-1), 0.5, WHITE),
    ]))
    els.append(mini_t)
    els.append(Spacer(1, 10))

    # Group by theme
    theme_order = ["Organizational Controls", "People Controls",
                   "Physical Controls", "Technological Controls"]
    themes_seen = {v["theme"] for v in soa.values()}
    for theme in theme_order:
        if theme not in themes_seen:
            continue
        theme_controls = [(cid, v) for cid, v in sorted(soa.items()) if v["theme"] == theme]
        els.append(Paragraph(f"<b>{theme}</b>", ParagraphStyle(
            name=f"th_{theme}", fontName="Helvetica-Bold", fontSize=9,
            textColor=DARK_BLUE, spaceBefore=8, spaceAfter=4)))

        headers = ["Control ID", "Control Name", "Applicability", "Implementation Status"]
        widths  = [60, pw*0.50, 80, 100]
        rows = [headers]
        for cid, v in theme_controls:
            rows.append([cid, v["name"], v["app"], v["stat"] or "—"])
        t = build_table(rows, widths, sty)
        # colour applicability column
        app_cmds = []
        for i, (_, v) in enumerate(theme_controls, start=1):
            if v["app"] == "Applicable":
                app_cmds += [("BACKGROUND", (2, i), (2, i), C_LOW),
                              ("TEXTCOLOR",  (2, i), (2, i), WHITE),
                              ("FONTNAME",   (2, i), (2, i), "Helvetica-Bold")]
            else:
                app_cmds += [("BACKGROUND", (2, i), (2, i), GREY_LIGHT)]
        if app_cmds:
            t.setStyle(TableStyle(app_cmds))
        els.append(KeepTogether(t))
        els.append(Spacer(1, 6))

    els.append(PageBreak())
    return els


# ─────────────────────────────────────────────────────────────────────────────
# Risk Treatment section
# ─────────────────────────────────────────────────────────────────────────────
def build_treatment_section(treatments, risks, styles_obj):
    sty = styles_obj
    els = []
    els += section_header("Risk Treatment Plan", styles_obj)

    pw = landscape(A4)[0] - 60
    headers = ["Risk", "Treatment", "Description", "Owner",
               "Target Date", "Status", "Inherent\nScore", "Residual\nScore",
               "Residual\nLevel", "Reduction"]
    widths  = [pw*0.20, 55, pw*0.24, pw*0.08, 55, 75, 45, 45, 50, 55]

    rows = [headers]
    for t in treatments:
        inh  = t.get("inherent_score", 0) or 0
        res  = t.get("residual_score", 0) or 0
        red  = inh - res
        pct  = f"{(red / inh * 100):.0f}%" if inh > 0 else "—"
        od   = " ⚠" if is_overdue(t["treatment_status"], t["target_date"]) else ""
        rows.append([
            t["risk_title"],
            t["treatment_option"],
            t["treatment_description"] or "",
            t["treatment_owner"] or "",
            str(t["target_date"] or ""),
            (t["treatment_status"] or "") + od,
            str(inh),
            str(res),
            t["residual_risk_level"] or "",
            pct,
        ])

    t_table = build_table(rows, widths, sty)
    t_table = severity_cell_style(t_table, rows, 8)
    els.append(t_table)
    els.append(Spacer(1, 8))
    return els


# ─────────────────────────────────────────────────────────────────────────────
# PDF builder
# ─────────────────────────────────────────────────────────────────────────────
def create_pdf_report(filename, scope, data):
    risks, treatments, mappings, assets, controls = data
    styles_obj = get_styles()
    gen_date   = datetime.now().strftime("%d %B %Y, %H:%M")

    doc = BaseDocTemplate(
        filename,
        pagesize=landscape(A4),
        rightMargin=30, leftMargin=30,
        topMargin=38, bottomMargin=30,
        title="Risk Assessment Framework — ISO/IEC 27001:2022-Aligned GRC Assessment",
        author="Areeb Amjad Khan",
        subject="Information Security Risk Assessment & GRC",
    )
    doc.report_scope = scope
    doc.gen_date = gen_date

    w, h = landscape(A4)
    # Define frames
    # Cover frame takes the whole page
    frame_cover = Frame(0, 0, w, h, id='cover_frame', 
                        topPadding=0, bottomPadding=0, rightPadding=0, leftPadding=0)
    # Normal frame has margins
    frame_normal = Frame(30, 30, w - 60, h - 68, id='normal_frame')

    def on_cover_page(c, doc_obj):
        draw_cover(c, doc_obj)

    template_cover = PageTemplate(id='Cover', frames=[frame_cover], onPage=on_cover_page)
    # Normal template uses NumberedCanvas logic later, so just standard frame here
    template_normal = PageTemplate(id='Normal', frames=[frame_normal])
    
    doc.addPageTemplates([template_cover, template_normal])

    elements = []

    # 1. Trigger the Cover Page
    # We add a blank paragraph to force the cover to render, then switch templates.
    elements.append(Paragraph(" ", styles_obj["Body"]))
    elements.append(NextPageTemplate('Normal'))
    elements.append(PageBreak())

    # 2. Exec Summary (always on Page 2)
    elements += build_exec_summary(risks, treatments, mappings, assets, styles_obj)

    # 3. Risk Matrix
    elements += build_matrix_section(risks, styles_obj)

    if scope in ("Full Risk Assessment", "Risk Register Only"):
        elements += build_top_risks(risks, styles_obj)
        elements += build_risk_register(risks, styles_obj)

    if scope in ("Full Risk Assessment", "Statement of Applicability"):
        elements += build_soa_section(mappings, styles_obj)

    if scope in ("Full Risk Assessment", "Risk Treatment Plan"):
        elements += build_treatment_section(treatments, risks, styles_obj)

    doc.build(elements, canvasmaker=NumberedCanvas)
    return filename


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────
# Excel palette (hex strings for openpyxl)
XL_NAVY      = "0D1B3E"
XL_DARK_BLUE = "1A3A6B"
XL_MID_BLUE  = "2E6DB4"
XL_LIGHT_BLUE= "D6E4F7"
XL_LIGHTEST  = "F2F6FC"
XL_GREY_LIGHT= "E8E8E8"
XL_GREY_MID  = "8C8C8C"

XL_CRITICAL  = "C0392B"
XL_HIGH      = "E67E22"
XL_MEDIUM    = "F1C40F"
XL_LOW       = "27AE60"
XL_WHITE     = "FFFFFF"

LEVEL_XL = {"Critical": XL_CRITICAL, "High": XL_HIGH,
             "Medium": XL_MEDIUM, "Low": XL_LOW}

STATUS_XL = {
    "Planned":          ("2E6DB4", XL_WHITE),
    "In Progress":      ("E67E22", XL_WHITE),
    "Implemented":      ("27AE60", XL_WHITE),
    "Accepted":         ("8C8C8C", XL_WHITE),
    "Closed":           ("0D1B3E", XL_WHITE),
}


def xl_header_style(fg=XL_NAVY, font_color=XL_WHITE, size=10):
    return {
        "font":      Font(bold=True, color=font_color, size=size),
        "fill":      PatternFill("solid", fgColor=fg),
        "alignment": Alignment(horizontal="center", vertical="center",
                                wrap_text=True),
        "border":    Border(
            bottom=Side(style="medium", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
        ),
    }


def xl_apply(cell, style_dict):
    for k, v in style_dict.items():
        setattr(cell, k, v)


def xl_set_col_widths(ws, widths):
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def xl_level_cell(cell, level):
    hex_c = LEVEL_XL.get(level)
    if hex_c:
        cell.fill = PatternFill("solid", fgColor=hex_c)
        cell.font = Font(bold=True, color=XL_WHITE, size=9)
        cell.alignment = Alignment(horizontal="center")


def xl_status_cell(cell, status):
    pair = STATUS_XL.get(status)
    if pair:
        bg, fg = pair
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(bold=True, color=fg, size=9)
        cell.alignment = Alignment(horizontal="center")


def xl_alt_row(ws, row, start_col, end_col, row_num):
    if row_num % 2 == 0:
        fill = PatternFill("solid", fgColor=XL_LIGHTEST)
        for c in range(start_col, end_col + 1):
            ws.cell(row=row, column=c).fill = fill


def create_excel_report(filename, scope, data):
    risks, treatments, mappings, assets, controls = data
    wb = Workbook()
    hdr_s = xl_header_style()

    # ── Executive Summary ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    kpis = compute_kpis(risks, treatments, mappings, assets)

    # Title block
    ws.merge_cells("A1:F1")
    ws["A1"] = "RISK ASSESSMENT FRAMEWORK — ISO/IEC 27001-ALIGNED GRC ASSESSMENT"
    ws["A1"].font = Font(bold=True, size=14, color=XL_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=XL_NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}  |  Prepared by: Areeb Amjad Khan  |  Organization: Sample Organization"
    ws["A2"].font = Font(italic=True, size=9, color=XL_GREY_MID)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])

    # KPI cards
    kpi_pairs = [
        ("Total Assets",          kpis["total_assets"],          XL_MID_BLUE),
        ("Total Risks",           kpis["total_risks"],            XL_DARK_BLUE),
        ("Critical Risks",        kpis["critical"],               XL_CRITICAL),
        ("High Risks",            kpis["high"],                   XL_HIGH),
        ("Medium Risks",          kpis["medium"],                 XL_MEDIUM),
        ("Low Risks",             kpis["low"],                    XL_LOW),
        ("Applicable Controls",   kpis["applicable_controls"],    XL_MID_BLUE),
        ("Risks w/o Treatment",   kpis["no_treatment"],           XL_HIGH),
        ("Open Treatments",       kpis["open_treatments"],        XL_DARK_BLUE),
        ("Overdue Treatments",    kpis["overdue_treatments"],     XL_CRITICAL),
    ]
    # Two rows of 5 KPIs
    for chunk_start in [0, 5]:
        chunk = kpi_pairs[chunk_start:chunk_start + 5]
        label_row = ws.max_row + 1
        val_row   = label_row + 1
        for col_i, (label, val, color) in enumerate(chunk, start=1):
            lc = ws.cell(row=label_row, column=col_i, value=label)
            lc.fill = PatternFill("solid", fgColor=color)
            lc.font = Font(bold=True, color=XL_WHITE, size=9)
            lc.alignment = Alignment(horizontal="center")
            vc = ws.cell(row=val_row, column=col_i, value=val)
            vc.fill = PatternFill("solid", fgColor=color)
            vc.font = Font(bold=True, color=XL_WHITE, size=18)
            vc.alignment = Alignment(horizontal="center")
            ws.row_dimensions[val_row].height = 30

    ws.append([])
    ws.append(["RISK DISTRIBUTION"])
    ws["A" + str(ws.max_row)].font = Font(bold=True, size=10, color=XL_DARK_BLUE)
    for level, count, color in [
        ("Critical", kpis["critical"], XL_CRITICAL),
        ("High",     kpis["high"],     XL_HIGH),
        ("Medium",   kpis["medium"],   XL_MEDIUM),
        ("Low",      kpis["low"],      XL_LOW),
    ]:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=level).font = Font(bold=True, color=XL_WHITE)
        ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=count).font = Font(bold=True)
    xl_set_col_widths(ws, [22, 14, 22, 14, 22, 14])
    ws.freeze_panes = "A3"

    # ── Asset Register ─────────────────────────────────────────────────────
    ws_a = wb.create_sheet("Asset Register")
    ws_a.sheet_view.showGridLines = False
    headers = ["ID", "Asset Name", "Type", "Description", "Owner", "Criticality", "Created At"]
    ws_a.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_a.cell(1, c), hdr_s)
    for i, a in enumerate(assets, 2):
        ws_a.append([a["id"], a["asset_name"], a["asset_type"],
                     a["description"], a["owner"], a["criticality"], a["created_at"]])
        xl_level_cell(ws_a.cell(i, 6), a["criticality"])
        xl_alt_row(ws_a, i, 1, 7, i)
    ws_a.freeze_panes = "A2"
    ws_a.auto_filter.ref = f"A1:G{ws_a.max_row}"
    xl_set_col_widths(ws_a, [5, 22, 14, 38, 18, 12, 18])

    # ── Risk Register ──────────────────────────────────────────────────────
    ws_r = wb.create_sheet("Risk Register")
    ws_r.sheet_view.showGridLines = False
    headers = ["ID", "Asset", "Risk Title", "Threat", "Vulnerability",
               "Existing Controls", "Likelihood", "Impact", "Risk Score",
               "Risk Level", "Risk Owner", "Status"]
    ws_r.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_r.cell(1, c), hdr_s)
    for i, r in enumerate(risks, 2):
        ws_r.append([r["id"], r["asset_name"], r["risk_title"], r["threat"],
                     r["vulnerability"], r["existing_controls"],
                     r["likelihood"], r["impact"], r["risk_score"],
                     r["risk_level"], r["risk_owner"], r["status"]])
        xl_level_cell(ws_r.cell(i, 10), r["risk_level"])
        xl_alt_row(ws_r, i, 1, 12, i)
    ws_r.freeze_panes = "A2"
    ws_r.auto_filter.ref = f"A1:L{ws_r.max_row}"
    xl_set_col_widths(ws_r, [5, 18, 30, 22, 22, 22, 10, 10, 10, 12, 16, 14])

    # ── Risk Treatment ─────────────────────────────────────────────────────
    ws_t = wb.create_sheet("Risk Treatment")
    ws_t.sheet_view.showGridLines = False
    headers = ["Risk Title", "Treatment", "Description", "Owner",
               "Target Date", "Status", "Inherent Score", "Inherent Level",
               "Residual Score", "Residual Level", "Reduction %"]
    ws_t.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_t.cell(1, c), hdr_s)
    for i, t in enumerate(treatments, 2):
        inh = t.get("inherent_score", 0) or 0
        res = t.get("residual_score", 0) or 0
        red_pct = f"{((inh - res) / inh * 100):.1f}%" if inh > 0 else "—"
        ws_t.append([
            t["risk_title"], t["treatment_option"], t["treatment_description"],
            t["treatment_owner"], t["target_date"], t["treatment_status"],
            inh, t.get("inherent_level", ""), res, t["residual_risk_level"], red_pct
        ])
        xl_level_cell(ws_t.cell(i, 8), t.get("inherent_level", ""))
        xl_level_cell(ws_t.cell(i, 10), t["residual_risk_level"] or "")
        xl_status_cell(ws_t.cell(i, 6), t["treatment_status"] or "")
        xl_alt_row(ws_t, i, 1, 11, i)
    ws_t.freeze_panes = "A2"
    ws_t.auto_filter.ref = f"A1:K{ws_t.max_row}"
    xl_set_col_widths(ws_t, [28, 12, 40, 16, 12, 14, 13, 13, 13, 13, 12])

    # ── ISO 27001 Controls ─────────────────────────────────────────────────
    ws_iso = wb.create_sheet("ISO 27001 Controls")
    ws_iso.sheet_view.showGridLines = False
    headers = ["Control ID", "Control Name", "Theme", "Description"]
    ws_iso.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_iso.cell(1, c), hdr_s)
    theme_colors = {
        "Organizational Controls": XL_DARK_BLUE,
        "People Controls":         XL_MID_BLUE,
        "Physical Controls":       "6B4226",
        "Technological Controls":  "1A6B4A",
    }
    for i, ctrl in enumerate(controls, 2):
        ws_iso.append([ctrl["control_id"], ctrl["control_name"],
                       ctrl["control_category"], ctrl["description"]])
        tc = theme_colors.get(ctrl["control_category"], XL_GREY_MID)
        ws_iso.cell(i, 3).fill = PatternFill("solid", fgColor=tc)
        ws_iso.cell(i, 3).font = Font(color=XL_WHITE, size=8)
        ws_iso.cell(i, 3).alignment = Alignment(horizontal="center")
        xl_alt_row(ws_iso, i, 1, 4, i)
    ws_iso.freeze_panes = "A2"
    ws_iso.auto_filter.ref = f"A1:D{ws_iso.max_row}"
    xl_set_col_widths(ws_iso, [12, 45, 24, 60])

    # ── Statement of Applicability ─────────────────────────────────────────
    ws_soa = wb.create_sheet("Statement of Applicability")
    ws_soa.sheet_view.showGridLines = False
    headers = ["Control ID", "Control Name", "Theme", "Applicability",
               "Justification", "Implementation Status"]
    ws_soa.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_soa.cell(1, c), hdr_s)
    soa_dict = {}
    for m in mappings:
        cid = m["control_id"]
        if cid not in soa_dict:
            soa_dict[cid] = dict(m)
        else:
            if m["applicability"] == "Applicable":
                soa_dict[cid]["applicability"] = "Applicable"
    for i, (cid, v) in enumerate(sorted(soa_dict.items()), 2):
        ws_soa.append([cid, v["control_name"], v["control_category"],
                       v["applicability"], v["justification"] or "",
                       v["implementation_status"] or ""])
        if v["applicability"] == "Applicable":
            ws_soa.cell(i, 4).fill = PatternFill("solid", fgColor=XL_LOW)
            ws_soa.cell(i, 4).font = Font(bold=True, color=XL_WHITE)
        ws_soa.cell(i, 4).alignment = Alignment(horizontal="center")
        xl_alt_row(ws_soa, i, 1, 6, i)
    ws_soa.freeze_panes = "A2"
    ws_soa.auto_filter.ref = f"A1:F{ws_soa.max_row}"
    xl_set_col_widths(ws_soa, [12, 42, 22, 15, 40, 22])

    # ── Risk-Control Mapping ───────────────────────────────────────────────
    ws_map = wb.create_sheet("Risk-Control Mapping")
    ws_map.sheet_view.showGridLines = False
    headers = ["Risk Title", "Control ID", "Control Name", "Theme",
               "Applicability", "Justification", "Implementation Status", "Notes"]
    ws_map.append(headers)
    for c, h in enumerate(headers, 1):
        xl_apply(ws_map.cell(1, c), hdr_s)
    for i, m in enumerate(mappings, 2):
        ws_map.append([m["risk_title"], m["control_id"], m["control_name"],
                       m["control_category"], m["applicability"],
                       m["justification"], m["implementation_status"],
                       m["implementation_notes"]])
        xl_alt_row(ws_map, i, 1, 8, i)
    ws_map.freeze_panes = "A2"
    ws_map.auto_filter.ref = f"A1:H{ws_map.max_row}"
    xl_set_col_widths(ws_map, [28, 10, 38, 22, 14, 38, 22, 30])

    # ── Global row heights ─────────────────────────────────────────────────
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and len(cell.value) > 60:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for rd in sheet.row_dimensions.values():
            if rd.height is None:
                rd.height = 15

    wb.save(filename)
    return filename


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────────────────────
def render_reports_page():
    if not has_permission("VIEW_REPORTS"):
        st.error("You do not have permission to view this page.")
        return

    st.title("📊 Reporting System")
    st.markdown("Generate professional PDF and Excel risk assessment reports using live data.")

    for key in ("generated_pdf_bytes", "generated_pdf_name",
                "generated_excel_bytes", "generated_excel_name"):
        if key not in st.session_state:
            st.session_state[key] = None

    with st.form("report_form"):
        col1, col2 = st.columns(2)
        with col1:
            scope = st.selectbox("Report Scope", [
                "Full Risk Assessment",
                "Risk Register Only",
                "Statement of Applicability",
                "Risk Treatment Plan",
            ])
            f_asset = st.selectbox(
                "Filter: Asset",
                ["All"] + [a["asset_name"] for a in
                           execute_query("SELECT asset_name FROM assets", fetch_all=True)],
            )
        with col2:
            f_level = st.selectbox("Filter: Risk Level",
                                   ["All", "Critical", "High", "Medium", "Low"])
            f_status = st.selectbox(
                "Filter: Risk Status",
                ["All"] + list(set(
                    r["status"] for r in
                    execute_query("SELECT status FROM risks", fetch_all=True)
                )),
            )
        btn_col1, btn_col2 = st.columns(2)
        gen_pdf = btn_col1.form_submit_button("📄 Generate PDF")
        gen_exc = btn_col2.form_submit_button("📊 Generate Excel")

    # Generation — outside form
    if gen_pdf:
        if not has_permission("GENERATE_REPORTS"):
            st.error("You do not have permission to generate reports.")
        else:
            try:
                data = fetch_report_data(scope, f_asset, f_level, f_status)
                ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn   = f"{REPORTS_DIR}/RAF_Report_{ts}.pdf"
                create_pdf_report(fn, scope, data)
                with open(fn, "rb") as fh:
                    st.session_state["generated_pdf_bytes"] = fh.read()
                st.session_state["generated_pdf_name"] = f"RAF_Report_{ts}.pdf"
                log_action("Report Generated", "Reports", f"Generated PDF report: {scope}", "reports", None)
            except Exception as e:
                st.error(f"PDF generation failed: {e}")

    if gen_exc:
        if not has_permission("GENERATE_REPORTS"):
            st.error("You do not have permission to generate reports.")
        else:
            try:
                data = fetch_report_data(scope, f_asset, f_level, f_status)
                ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
                fn   = f"{REPORTS_DIR}/RAF_Report_{ts}.xlsx"
                create_excel_report(fn, scope, data)
                with open(fn, "rb") as fh:
                    st.session_state["generated_excel_bytes"] = fh.read()
                st.session_state["generated_excel_name"] = f"RAF_Report_{ts}.xlsx"
                log_action("Report Generated", "Reports", f"Generated Excel report: {scope}", "reports", None)
            except Exception as e:
                st.error(f"Excel generation failed: {e}")

    # Download buttons — outside form
    if st.session_state["generated_pdf_bytes"] is not None:
        st.success(f"✅ PDF ready: {st.session_state['generated_pdf_name']}")
        st.download_button(
            label="⬇️ Download PDF",
            data=st.session_state["generated_pdf_bytes"],
            file_name=st.session_state["generated_pdf_name"],
            mime="application/pdf",
            key="dl_pdf",
        )

    if st.session_state["generated_excel_bytes"] is not None:
        st.success(f"✅ Excel ready: {st.session_state['generated_excel_name']}")
        st.download_button(
            label="⬇️ Download Excel",
            data=st.session_state["generated_excel_bytes"],
            file_name=st.session_state["generated_excel_name"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel",
        )
